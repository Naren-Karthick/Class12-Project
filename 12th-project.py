import mysql.connector as sqltor
import customtkinter as tk
from tkinter import ttk
from openpyxl import *
import tkinter.messagebox as tkmb
from tkinter import *


def student():

    studwin = tk.CTk()
    studwin.title("Student Database Management")
    studwin.geometry("1300x700")
    studwin.state("zoomed")

    studtab = tk.CTkTabview(studwin, width=1250, height=650)
    studtab.pack(pady=20, padx=20)

    displaytab = studtab.add("Display")
    inserttab = studtab.add("Insert")
    updatetab = studtab.add("Update")
    deletetab = studtab.add("Delete")
    getexceltab = studtab.add("Get Excel")

    def change(event):
        for j in tree.get_children():
            tree.delete(j)
        inval = getall(combo.get())
        for i in range(len(inval)):
            tree.insert(parent="", index="end", iid=i, values=inval[i])

    def changein():
        for j in tree.get_children():
            tree.delete(j)
        inval1 = getall(combo.get())
        for i in range(len(inval1)):
            tree.insert(parent="", index="end", iid=i, values=inval1[i])

    def getall(x):
        cur = con.cursor()

        if x == "All":
            sql = f"select * from student"
        else:
            sql = f"select * from student where class_sec = '{x}'"
        cur.execute(sql)
        return cur.fetchall()

    def getad():
        cur = con.cursor()
        sql = "Select admn_id from student"
        cur.execute(sql)
        return cur.fetchall()

    def getval():
        res = []
        for i in range(1, 13):
            for j in range(65, 70):
                res.append(f"{i}" f"{chr(j)}")
        return res

    lab1 = tk.CTkLabel(displaytab, text="Select the class_sec", font=("Arial", 18))
    lab1.pack()
    val = getval()
    val.insert(0, "All")
    combo = ttk.Combobox(displaytab, width=20, values=val, state="readonly")
    combo.pack(pady=10)
    combo.bind("<<ComboboxSelected>>", change)
    combo.set("All")

    tree = ttk.Treeview(displaytab, height=600)
    tree.pack(padx=10, pady=10)
    tree["columns"] = (
        "Admn_ID",
        "Name",
        "class_sec",
        "Dob",
        "Father Name",
        "Mother Name",
    )
    tree.column("#0", width=0)
    tree.column("Admn_ID", width=70, anchor="center", minwidth=30)
    tree.column("Name", width=160, anchor="center")
    tree.column("class_sec", width=80, anchor="center")
    tree.column("Dob", width=120, anchor="center")
    tree.column("Father Name", width=160, anchor="center")
    tree.column("Mother Name", width=160, anchor="center")

    tree.heading("Admn_ID", text="Admn_Id")
    tree.heading("Name", text="Name")
    tree.heading("class_sec", text="Class_Sec")
    tree.heading("Dob", text="Dob")
    tree.heading("Father Name", text="Father_Name")
    tree.heading("Mother Name", text="Mother_Name")
    change("All")

    # Insert tab
    inframe = tk.CTkFrame(inserttab)
    inframe.pack()
    adidlab = tk.CTkLabel(inframe, text="Admn_Id", font=("Arial", 24))
    adidlab.grid(row=0, column=0, padx=20, pady=30)
    adiden = tk.CTkEntry(
        inframe,
        placeholder_text="Enter the Admission ID of Student",
        width=210,
        height=32,
    )
    adiden.grid(row=0, column=1, padx=20, pady=30)

    namelab = tk.CTkLabel(inframe, text="Name", font=("Arial", 24))
    namelab.grid(row=0, column=3, padx=20, pady=30)
    namen = tk.CTkEntry(
        inframe, placeholder_text="Enter the Name of Student", width=200, height=32
    )
    namen.grid(row=0, column=4, padx=20, pady=30)

    class_seclab = tk.CTkLabel(inframe, text="Class_Sec", font=("Arial", 24))
    class_seclab.grid(row=1, column=0, padx=20, pady=30)
    classcombo = ttk.Combobox(inframe, width=20, values=getval(), state="readonly")
    classcombo.grid(row=1, column=1, padx=20, pady=30)

    doblab = tk.CTkLabel(inframe, text="Dob", font=("Arial", 24))
    doblab.grid(row=1, column=3, padx=20, pady=30)
    doben = tk.CTkEntry(
        inframe, placeholder_text="Enter the Dob of Student", width=200, height=32
    )
    doben.grid(row=1, column=4, padx=20, pady=30)

    fnamelab = tk.CTkLabel(inframe, text="Father Name", font=("Arial", 24))
    fnamelab.grid(row=2, column=0, padx=20, pady=30)
    fnamen = tk.CTkEntry(
        inframe,
        placeholder_text="Enter the Father Name of Student",
        width=200,
        height=32,
    )
    fnamen.grid(row=2, column=1, padx=20, pady=30)

    mnamelab = tk.CTkLabel(inframe, text="Mother Name", font=("Arial", 24))
    mnamelab.grid(row=2, column=3, padx=20, pady=30)
    mnamen = tk.CTkEntry(
        inframe,
        placeholder_text="Enter the Mother Name of Student",
        width=210,
        height=32,
    )
    mnamen.grid(row=2, column=4, padx=20, pady=30)

    def submit():
        admn = adiden.get() or "NULL"
        name = namen.get() or "NULL"
        class_sec = classcombo.get() or "NULL"
        dob = doben.get() or "NULL"
        fname = fnamen.get() or "NUll"
        mname = mnamen.get() or "NULL"
        cur = con.cursor()
        sql = f"insert into student values({admn},'{name}','{class_sec}','{dob}','{fname}','{mname}')"
        cur.execute(sql)
        con.commit()
        adiden.delete(0, END)
        namen.delete(0, END)
        doben.delete(0, END)
        fnamen.delete(0, END)
        mnamen.delete(0, END)
        tkmb.showinfo("Insert", "Inserted Succesfully")
        studtab.set("Display")

        changein()

    getbut = tk.CTkButton(inserttab, text="Submit", command=submit)
    getbut.pack(pady=30)
    msg = tk.CTkLabel(inserttab, text="", font=("Arial", 30))
    msg.pack()

    # update tab

    upframe = tk.CTkFrame(updatetab)
    upframe.pack()

    admnidlab = tk.CTkLabel(upframe, text="Admn_Id", font=("Arial", 24))
    admnidlab.grid(row=0, column=0, padx=20, pady=30)
    admniden = tk.CTkEntry(
        upframe,
        placeholder_text="Enter the Admission ID of Student",
        width=210,
        height=32,
    )
    admniden.grid(row=0, column=1, padx=50, pady=30)

    label1 = tk.CTkLabel(upframe, text="Select the Field:", font=("Arial", 24))
    label1.grid(row=1, column=0, padx=20, pady=30)

    lst = ["Name", "Class_Sec", "Dob", "Father_Name", "Mother_Name"]
    lstoption = tk.CTkOptionMenu(
        upframe,
        values=lst,
        width=160,
        anchor="center",
    )
    lstoption.grid(row=2, column=0, padx=10, pady=5)

    label2 = tk.CTkLabel(upframe, text="Enter the Value to change:", font=("Arial", 24))
    label2.grid(row=1, column=1, padx=20, pady=20)

    chen = tk.CTkEntry(
        upframe, placeholder_text="Value Here", width=200, font=("Arial", 20)
    )
    chen.grid(row=2, column=1, padx=20)

    def update():
        f = lstoption.get()
        a = admniden.get()
        v = chen.get()
        if f == "Father_Name":
            f = "fname"
        if f == "Mother_Name":
            f = "mname"
        cur = con.cursor()
        admnidlst = getad()
        if a in admnidlst:
            sql = f"update student set {f} = '{v}' where admn_id = {a}"
            cur.execute(sql)
            con.commit()
            admniden.delete(0, END)
            chen.delete(0, END)
            tkmb.showinfo("Updated", "Updated Succesfully")
            studtab.set("Display")
            changein()
        else:
            tkmb.showerror("Warning", "INVAID ADMN_ID")

    subbut = tk.CTkButton(updatetab, text="Submit", command=update)
    subbut.pack(pady=30)

    # delete tab

    delframe = tk.CTkFrame(deletetab)
    delframe.pack()

    admndelid = tk.CTkLabel(
        delframe, text="Enter the Admission ID to delete", font=("Arial", 26)
    )
    admndelid.grid(row=0, column=0, padx=20, pady=30)

    admndelen = tk.CTkEntry(
        delframe, width=210, placeholder_text="Value Here", font=("Arial", 20)
    )
    admndelen.grid(row=1, column=0, padx=20, pady=30)

    def delete():
        cur = con.cursor()
        x = admndelen.get()
        admnid = getad()
        if x in admnid:
            sql = f"delete from student where admn_ID = {x}"
            cur.execute(sql)
            con.commit()
            admndelen.delete(0, END)
            tkmb.showinfo("delete", "Deleted Succesfully")
            studtab.set("Display")
            changein()
        else:
            tkmb.showerror("Warning", "Invaid Admn_Id")

    delbut = tk.CTkButton(delframe, text="Submit", width=200, command=delete)
    delbut.grid(row=2, column=0, padx=20, pady=10)

    # get excel tab

    getframe = tk.CTkFrame(getexceltab)
    getframe.pack()

    getlab = tk.CTkLabel(getframe, text="Select Class_Sec:", font=("Arial", 26))
    getlab.grid(row=0, column=0, padx=20, pady=30)

    """def class_sec():
        cur = con.cursor()
        sql = "select distinct class_sec from student order by class_sec"
        cur.execute(sql)
        return cur.fetchall()"""

    getcombo = ttk.Combobox(getframe, values=getval(), state="readonly")
    getcombo.grid(row=1, column=0, padx=10, pady=20)

    def getex():
        cur = con.cursor()
        x = getcombo.get()
        sql = f"select * from student where class_sec='{x}'"
        cur.execute(sql)
        y = cur.fetchall()

        try:
            wb = load_workbook(filename=f"{x}.xlsx")
        except:
            wb = Workbook()
        sheet = wb.active
        i = 0
        for row in y:
            i += 1
            j = 1
            for col in row:
                cell = sheet.cell(row=i, column=j)
                cell.value = col
                j += 1

        wb.save(filename=f"{x}.xlsx")
        tkmb.showinfo("Getexcel", "Extracted Succesfully")
        studtab.set("Display")

    getbut = tk.CTkButton(getframe, text="Sumbit", width=200, command=getex)
    getbut.grid(row=2, column=0, padx=20, pady=10)

    studwin.mainloop()


def connect():
    mycon = sqltor.connect(host="localhost", user="root", password="Narenguru2007")
    mycur = mycon.cursor()
    mycur.execute("create database if not exists naren_12th_project")
    mycur.execute("use naren_12th_project")
    mycur.execute(
        "create table if not exists student(admn_id int primary key , name varchar(30), class_sec varchar(10),dob date,fname varchar(30),mname varchar(30))"
    )
    mycon.commit()
    mycon.close()


if __name__ == "__main__":
    tk.set_appearance_mode("dark")
    tk.set_default_color_theme("blue")

    connect()

    con = sqltor.connect(
        host="localhost",
        user="root",
        password="Narenguru2007",
        database="naren_12th_project",
    )
    student()
